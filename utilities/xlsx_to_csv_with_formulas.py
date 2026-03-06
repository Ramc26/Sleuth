"""
utilities/xlsx_to_csv_with_formulas.py
========================================
Converts each tab of SamplePaysheet.xlsx to a separate CSV file.
Each CSV contains:
  1. The raw data rows (computed values, not formulas)
  2. A FORMULA REFERENCE section appended below with a clear separator.
     This section documents every column's Excel formula and its
     plain-English explanation — used as metadata for Sleuth's
     payroll engine to understand and reproduce the calculations.

Usage:
    python3 utilities/xlsx_to_csv_with_formulas.py

Output files (written to data/payroll_reference/):
    Anchors_Feb26.csv         — per-employee payroll output sheet
    Anchor_Attendance.csv     — monthly attendance input sheet
    Jan26_LeaveBal.csv        — previous month closing leave balances

Run from the project root:
    cd /home/ramarao/Desktop/misc/Sleuth
    python3 utilities/xlsx_to_csv_with_formulas.py
"""

import csv
import os
import re
import textwrap
from pathlib import Path

import openpyxl

# ── Config ─────────────────────────────────────────────────────────────────
XLSX_PATH   = Path(__file__).parent.parent / "SamplePaysheet.xlsx"
OUTPUT_DIR  = Path(__file__).parent.parent / "data" / "payroll_reference"
SEPARATOR   = "## FORMULA_REFERENCE_START"
END_SEP     = "## FORMULA_REFERENCE_END"

# ── Human-readable formula explanations ────────────────────────────────────
# Key: column header name (strip spaces)  →  (formula_template, plain_english)
ANCHORS_FORMULAS = {
    "DOE": (
        "=VLOOKUP(EMP_ID, Attendance!B:K, 10, 0)",
        "Date of Exit — pulled from Attendance sheet column K (DOE). "
        "If no resignation, cell is manually set to 'No'.",
    ),
    "PayableDays": (
        "=VLOOKUP(EMP_ID, Attendance!B:BB, 53, 0)",
        "Pulled from Attendance sheet column BB (Total Payable Days). "
        "Formula there: Total_Days − (UL + Leave_Balance_LOP). "
        "The constant 28 is the divisor (stored in cell O1).",
    ),
    "StandedMonthSalary": (
        "= FIXED (12360 for Anchor, 17000 for AMIA/Asset)",
        "Fixed slab salary for the employee category. "
        "Anchor = ₹12,360; AMIA / Asset = ₹17,000.",
    ),
    "CurrentMonthSalary": (
        "= StdSalary / 28 * PayableDays",
        "Prorated salary. Divisor is always 28 (Feb month days, "
        "stored in cell O1). Rounded to whole rupee by downstream formulas.",
    ),
    "Basic": (
        "= CurrentMonthSalary * 100%   [Anchor]"
        "\n= CurrentMonthSalary * 90%   [AMIA/Asset]",
        "Basic component of salary. For standard Anchors = 100% of current salary. "
        "For AMIA/Asset = 90% (HRA is the remaining 10%).",
    ),
    "HRA": (
        "= 0                            [Anchor]"
        "\n= CurrentMonthSalary * 10%   [AMIA/Asset]",
        "House Rent Allowance. Zero for standard Anchors. "
        "10% of current month salary for AMIA/Asset employees.",
    ),
    "LTA": (
        "= 0  (all categories)",
        "Leave Travel Allowance — currently zero for all employees.",
    ),
    "SpecialAllowances": (
        "= 0  (all categories)",
        "Special Allowances — currently zero for all employees.",
    ),
    "GrossSalary": (
        "= ROUND(Basic + HRA + LTA + SpecialAllowances, 0)",
        "Gross Salary = sum of all salary components, rounded to nearest rupee. "
        "For standard Anchors this equals CurrentMonthSalary. "
        "For AMIA/Asset: Basic (90%) + HRA (10%) = CurrentMonthSalary.",
    ),
    "OneTimeBonus": (
        "= 12360 − GrossSalary          [if resigned & no bonus exclusion]"
        "\n= 0                            [active employees]"
        "\n= 0                            [resigned with exclusion flag]",
        "Top-up payment for resigned employees who worked most of the month "
        "(typically DOE ≥ 15th). Makes their total gross equal to the full "
        "standard month salary (₹12,360 for Anchors). "
        "Manually overridden to 0 where HR decides not to pay.",
    ),
    "Gratuity": (
        "= GratuityAmount  (cell AR, calculated separately)",
        "Statutory gratuity for employees with ≥ 5 completed years of service. "
        "Formula: ROUND(LastDrawnBasic × 15 / 26 × CompletedYears, 0). "
        "Zero for active employees or resigned with < 5 years.",
    ),
    "FinalGross": (
        "= GrossSalary + OneTimeBonus + Gratuity",
        "Total gross payment: salary + bonus (if any) + gratuity (if any).",
    ),
    "GrossForPF": (
        "= ROUND(Basic + LTA + SpecialAllowances − HRA, 0)"
        "\n= ROUND(GrossSalary − HRA, 0)",
        "PF-eligible gross: Gross minus HRA (HRA is excluded from PF basis). "
        "For standard Anchors = GrossSalary (since HRA = 0).",
    ),
    "EPFWages": (
        "= ROUND(IF(GrossForPF > 15000, 15000, GrossForPF), 0)",
        "EPF contribution ceiling: capped at ₹15,000. "
        "If PF-eligible gross exceeds ₹15,000, EPF is calculated on ₹15,000 only.",
    ),
    "PF12pct": (
        "= ROUND(IF(GrossForPF > 15000, 1800, GrossForPF × 12%), 0)",
        "Employee EPF contribution at 12%. Capped at ₹1,800 (12% of ₹15,000) "
        "when gross exceeds the ceiling.",
    ),
    "VPF": (
        "= 0  (Voluntary PF — not applicable currently)",
        "Voluntary Provident Fund — zero for all employees.",
    ),
    "ESI075pct": (
        "= ROUNDUP(IF(StdSalary > 21000, 0, (FinalGross − Gratuity) × 0.75%), 0)",
        "Employee ESI contribution at 0.75%. Applied on Gross (excl. gratuity). "
        "Zero if standard salary > ₹21,000 (ESIC exemption threshold). "
        "Uses ROUNDUP (rounds fractions up, not to nearest).",
    ),
    "EPF833pct": (
        "= ROUND(PF12pct / 12% × 8.33%, 0)"
        "\n= ROUND(PF12pct × (8.33 / 12), 0)",
        "Employer Pension contribution (EPS) at 8.33% of EPF wages. "
        "Derived from employee PF amount: employee_PF ÷ 12% × 8.33%. "
        "Capped because PF is already capped at ₹15,000 ceiling.",
    ),
    "PF367pct": (
        "= PF12pct − EPF833pct",
        "Employer PF contribution at 3.67% (= 12% employee − 8.33% pension). "
        "The remaining employer share after pension allocation.",
    ),
    "ESI325pct": (
        "= IF(StdSalary > 21000, 0, (FinalGross − Gratuity) × 3.25%)",
        "Employer ESI contribution at 3.25% on Gross (excl. gratuity). "
        "Zero if standard salary exceeds ₹21,000 ESIC exemption limit. "
        "Note: NOT rounded (unlike employee ESI which uses ROUNDUP).",
    ),
    "TDS": (
        "= 0  (manual entry where applicable)",
        "Tax Deducted at Source — entered manually for employees with TDS liability.",
    ),
    "AnyOthersAdvance": (
        "= 0  (manual entry)",
        "Any other deductions or advances — entered manually.",
    ),
    "InsuranceAdvance": (
        "= 0  (manual entry)",
        "Insurance advance recovery — entered manually.",
    ),
    "ProfessionTax": (
        "= IF((FinalGross − Gratuity) >= 20001, 200,"
        "\n    IF((FinalGross − Gratuity) >= 15001, 150, 0))",
        "State Profession Tax (Karnataka/Telangana slab): "
        "₹200 if gross ≥ ₹20,001; ₹150 if gross ≥ ₹15,001; ₹0 otherwise. "
        "Applied on Gross excluding gratuity.",
    ),
    "LWF": (
        "= 0  (Labour Welfare Fund — not applicable currently)",
        "Labour Welfare Fund deduction — zero for all employees currently.",
    ),
    "PF_ESI_LWF_Total": (
        "= PF12pct + ESI075pct + LWF",
        "Total of statutory employee deductions: PF + ESI + LWF.",
    ),
    "NetSalary": (
        "= SUM(ROUND(FinalGross − (TDS + Others + Insurance + ProfTax + PF_ESI_LWF), 0), 0)",
        "Net take-home salary: Final Gross minus all deductions. "
        "Rounded to nearest rupee.",
    ),
    "CompletedYears": (
        "= ROUND((DOE − DOJ) / 365, 0)",
        "Completed years of service, rounded to nearest year. "
        "ROUND(..., 0) means ≥ 6 months of a year counts as a full year "
        "(Indian Gratuity Act). Used only for resigned employees.",
    ),
    "LastDrawnBasic": (
        "= StdSalary  (standard monthly salary)",
        "Last drawn basic salary — taken as the standard monthly salary "
        "(₹12,360 for Anchor; ₹15,300 for AMIA/Asset).",
    ),
    "GratuityAmount": (
        "= ROUND(LastDrawnBasic × 15 / 26 × CompletedYears, 0)",
        "Statutory gratuity: (Basic ÷ 26) × 15 × Years. "
        "Only for resigned employees with ≥ 5 completed years. "
        "₹0 for active employees or those with < 5 years.",
    ),
}

ATTENDANCE_FORMULAS = {
    "TotalPresent": (
        "= COUNTIF(Day1:Day28, 'P') + COUNTIF(Day1:Day28, 'HFL') / 2",
        "Count of Present days. Half-Day Leaves count as 0.5 present.",
    ),
    "TotalWeekOffs": (
        "= COUNTIF(Day1:Day28, 'W/O')",
        "Count of Week Off days in the month.",
    ),
    "TotalLeaves": (
        "= COUNTIF(Day1:Day28, 'L')",
        "Count of Paid Leave days taken.",
    ),
    "TotalHalfDayLeaves": (
        "= COUNTIF(Day1:Day28, 'HFL') / 2",
        "Half-Day Leaves converted to full-day equivalent.",
    ),
    "Holiday": (
        "= COUNTIF(Day1:Day28, 'H')",
        "Count of public Holiday days.",
    ),
    "Marriage": (
        "= COUNTIF(Day1:Day28, 'MG')",
        "Count of Marriage Leave days.",
    ),
    "MaternityLeave": (
        "= COUNTIF(Day1:Day28, 'ML')",
        "Count of Maternity Leave days.",
    ),
    "UnPaidLeave": (
        "= COUNTIF(Day1:Day28, 'UL')",
        "Count of Unpaid Leave days. These directly reduce payable days.",
    ),
    "Miscarriage": (
        "= COUNTIF(Day1:Day28, 'MC')",
        "Count of Miscarriage Leave days.",
    ),
    "BereavementLeaves": (
        "= COUNTIF(Day1:Day28, 'BL')",
        "Count of Bereavement Leave days.",
    ),
    "MedicalLeave": (
        "= COUNTIF(Day1:Day28, 'MDL')",
        "Count of Medical Leave days.",
    ),
    "LossOfPay_Att": (
        "= COUNTIF(Day1:Day28, 'LOP')",
        "Count of days marked LOP directly in attendance (rare). "
        "The actual LOP for salary deduction is calculated in the Leave Balance section.",
    ),
    "TotalDays": (
        "= SUM(TotalPresent : LossOfPay_Att)",
        "Sum of ALL attendance categories = total calendar days tracked "
        "(= 28 for full month, or days up to DOE for resigned employees).",
    ),
    "TotalActualDays": (
        "= SUM(TotalLeaves, HalfDayLeaves, Marriage, ML, UL, MC, BL, MDL, LOP, TotalPresent)",
        "Working days excluding Week Offs and Holidays. "
        "Used as denominator for Attendance %.",
    ),
    "TotalPayableDays": (
        "= TotalDays − (UnPaidLeave + LeaveBalance_LOP)",
        "KEY: Days for salary calculation. "
        "TotalDays minus UL (unpaid) minus LOP from leave balance utilization. "
        "LOP (column BO) is computed in the leave balance utilization section.",
    ),
    "TotalPct": (
        "= TotalPresent / TotalActualDays",
        "Attendance percentage = Present days ÷ working days.",
    ),
    "Opening_LastYearCF": (
        "= VLOOKUP(EMP_ID, Jan26_LeaveBal!B:D, 3, 0)",
        "Opening carry-forward leaves from previous month's closing balance.",
    ),
    "Opening_CL": (
        "= VLOOKUP(EMP_ID, Jan26_LeaveBal!B:E, 4, 0) + 0.5",
        "Opening Casual Leave = January closing CL + monthly accrual (0.5/month).",
    ),
    "Opening_SL": (
        "= VLOOKUP(EMP_ID, Jan26_LeaveBal!B:F, 5, 0) + 0.5",
        "Opening Sick Leave = January closing SL + monthly accrual (0.5/month).",
    ),
    "Opening_EL": (
        "= VLOOKUP(EMP_ID, Jan26_LeaveBal!B:G, 6, 0) + 1.0",
        "Opening Earned Leave = January closing EL + monthly accrual (1.0/month).",
    ),
    "Opening_MG": (
        "= VLOOKUP(EMP_ID, Jan26_LeaveBal!B:H, 7, 0)",
        "Opening Marriage Leave — no monthly accrual; carries forward as-is.",
    ),
    "Opening_ExtraEL": (
        "= VLOOKUP(EMP_ID, Jan26_LeaveBal!B:I, 8, 0) + 0.25",
        "Opening Extra EL = January closing Extra EL + monthly accrual (0.25/month).",
    ),
    "Util_LastYearCF": (
        "= IF((L + HFL) <= Opening_CF, (L + HFL), Opening_CF)",
        "Leave utilization drawn from Carry-Forward balance first.",
    ),
    "Util_CL": (
        "= IF(((L + HFL) − Util_CF) > Opening_CL, Opening_CL, (L + HFL) − Util_CF)",
        "Casual Leave utilization (after exhausting CF balance).",
    ),
    "Util_SL": (
        "= IF((L + HFL − Util_CF − Util_CL) >= Opening_SL, Opening_SL, L + HFL − Util_CF − Util_CL)",
        "Sick Leave utilization (after exhausting CF and CL).",
    ),
    "Util_EL": (
        "= IF((L + HFL − Util_CF − Util_CL − Util_SL) >= Opening_EL, Opening_EL,"
        "\n     L + HFL − Util_CF − Util_CL − Util_SL)",
        "Earned Leave utilization (after exhausting CF, CL, SL).",
    ),
    "Util_MG": (
        "= Marriage_days  (column AS)",
        "Marriage Leave utilization = number of MG days taken this month.",
    ),
    "Util_LOP": (
        "= (L + HFL) − Util_CF − Util_CL − Util_SL − Util_EL",
        "KEY — Loss of Pay: leave days that exceeded ALL available balances. "
        "This value feeds into TotalPayableDays = TotalDays − UL − LOP.",
    ),
    "Closing_CF": (
        "= Opening_CF − Util_CF",
        "Closing carry-forward leave balance.",
    ),
    "Closing_CL": (
        "= Opening_CL − Util_CL",
        "Closing Casual Leave balance (carry to next month).",
    ),
    "Closing_SL": (
        "= Opening_SL − Util_SL",
        "Closing Sick Leave balance.",
    ),
    "Closing_EL": (
        "= Opening_EL − Util_EL",
        "Closing Earned Leave balance.",
    ),
    "Closing_MG": (
        "= Opening_MG − Util_MG",
        "Closing Marriage Leave balance.",
    ),
    "Closing_ExtraEL": (
        "= Opening_ExtraEL  (no deduction — Extra EL accumulates)",
        "Closing Extra EL = same as Opening (not consumed this month).",
    ),
}

LEAVE_BAL_FORMULAS = {
    "LastYearCF": (
        "= manual / carried from previous payroll run",
        "Last Year carry-forward leave balance (rarely non-zero for active employees).",
    ),
    "CL": (
        "= closing CL from previous month's attendance sheet",
        "Closing Casual Leave balance from the prior month. "
        "Used as opening balance for next month (+ 0.5 accrual).",
    ),
    "SL": (
        "= closing SL from previous month's attendance sheet",
        "Closing Sick Leave balance from the prior month. "
        "Used as opening balance for next month (+ 0.5 accrual).",
    ),
    "EL": (
        "= closing EL from previous month's attendance sheet",
        "Closing Earned Leave balance from the prior month. "
        "Used as opening balance for next month (+ 1.0 accrual).",
    ),
    "MG": (
        "= closing MG from previous month's attendance sheet",
        "Closing Marriage Leave balance. No accrual.",
    ),
    "ExtraELBalance": (
        "= closing Extra EL from previous month's attendance sheet",
        "Closing Extra EL balance from prior month (+ 0.25 accrual next month).",
    ),
}


# ── Helpers ─────────────────────────────────────────────────────────────────
def col_letter(n: int) -> str:
    """Convert 1-based column index to Excel letter (A, B, ..., Z, AA, ...)."""
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def get_sheet_data(ws_values, ws_formulas):
    """
    Return (headers_list, data_rows, formula_rows) for a worksheet.
    headers_list: list of (col_idx, col_letter, header_value)
    data_rows: list of row-tuples (values)
    formula_rows: same structure but with raw cell formulas
    """
    rows_val = list(ws_values.iter_rows(values_only=True))
    rows_fml = list(ws_formulas.iter_rows(values_only=True))
    return rows_val, rows_fml


def extract_first_formula_row(ws_formulas, data_start_row: int) -> dict:
    """
    Extract formulas from the first data row.
    Returns {col_letter: formula_string}
    """
    formulas = {}
    row = list(ws_formulas.iter_rows(
        min_row=data_start_row,
        max_row=data_start_row,
        values_only=False
    ))[0]
    for cell in row:
        val = str(cell.value) if cell.value else ""
        if val.startswith("="):
            formulas[cell.column_letter] = val
    return formulas


def generalise_formula(formula: str, row_num: int) -> str:
    """
    Replace row numbers that appear as part of cell references (e.g. A3, AB3)
    with 'N', while leaving numbers in constants alone (365, 15000, 12%, etc.).
    """
    # Match: one or more uppercase letters followed by the row number
    # e.g. A3 → AN, AB3 → ABN, $A$3 → $A$N
    return re.sub(r'(?<=[A-Z])' + str(row_num) + r'(?!\d)', 'N', formula)


def write_csv_with_formulas(
    output_path: Path,
    sheet_name: str,
    ws_values,
    ws_formulas,
    header_row: int,        # 1-based row containing column headers
    data_start: int,        # 1-based first data row
    formula_dict: dict,     # our manual formula docs
):
    print(f"  Writing {output_path.name} ...")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    rows_val = list(ws_values.iter_rows(values_only=True))
    raw_formulas = extract_first_formula_row(ws_formulas, data_start)

    # Build header from header_row (0-indexed)
    header = list(rows_val[header_row - 1])
    # Data rows
    data = rows_val[data_start - 1:]

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)

        # ── 1. Sheet metadata ─────────────────────────────────────────────
        writer.writerow([f"## SHEET: {sheet_name}"])
        writer.writerow([f"## SOURCE: SamplePaysheet.xlsx"])
        writer.writerow([f"## Generated by: utilities/xlsx_to_csv_with_formulas.py"])
        writer.writerow([])

        # ── 2. Main header row ────────────────────────────────────────────
        writer.writerow(header)

        # ── 3. Data rows ──────────────────────────────────────────────────
        empty_count = 0
        for row in data:
            if all(v is None or str(v).strip() == "" for v in row):
                empty_count += 1
                if empty_count > 3:
                    break   # stop after several empty rows (summary rows, etc.)
                continue
            empty_count = 0
            # Convert dates to clean strings
            clean = []
            for v in row:
                if v is None:
                    clean.append("")
                elif hasattr(v, "strftime"):
                    clean.append(v.strftime("%Y-%m-%d"))
                else:
                    clean.append(v)
            writer.writerow(clean)

        # ── 4. Formula Reference section ──────────────────────────────────
        writer.writerow([])
        writer.writerow([SEPARATOR])
        writer.writerow([
            "COLUMN_LETTER", "COLUMN_HEADER",
            "EXCEL_FORMULA (row N = first data row)",
            "PLAIN_ENGLISH_EXPLANATION",
        ])

        # Build a map: header_name → col_letter
        header_to_col = {}
        for i, h in enumerate(header):
            if h:
                header_to_col[str(h).strip()] = col_letter(i + 1)

        # Write formulas from Excel (raw, from the first data row)
        written_cols = set()
        for col_ltr, fml in sorted(raw_formulas.items()):
            col_idx = ws_formulas[f"{col_ltr}1"].column - 1
            col_hdr = header[col_idx] if col_idx < len(header) else ""
            gen_fml = generalise_formula(fml, data_start)
            writer.writerow([col_ltr, col_hdr, gen_fml, ""])
            written_cols.add(col_ltr)

        # Write our manual formula docs (for columns without Excel formulas / hardcoded)
        writer.writerow([])
        writer.writerow(["## DETAILED FORMULA EXPLANATIONS"])
        writer.writerow(["COLUMN_KEY", "FORMULA_TEMPLATE", "PLAIN_ENGLISH_EXPLANATION"])
        for col_key, (fml_template, explanation) in formula_dict.items():
            writer.writerow([col_key, fml_template, explanation])

        writer.writerow([END_SEP])

    print(f"    → {output_path} ({output_path.stat().st_size:,} bytes)")


# ── Main ────────────────────────────────────────────────────────────────────
def main():
    print(f"Loading {XLSX_PATH} ...")
    wb_val = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    wb_fml = openpyxl.load_workbook(XLSX_PATH, data_only=False)

    print(f"Sheets found: {wb_val.sheetnames}\n")

    # ── 1. Anchors_Feb'26 ──────────────────────────────────────────────────
    # Row 1 = group headers (EMPLOYEE CONTRIBUTION, etc.)
    # Row 2 = column headers
    # Row 3 = first data row
    write_csv_with_formulas(
        output_path=OUTPUT_DIR / "Anchors_Feb26.csv",
        sheet_name="Anchors_Feb'26",
        ws_values=wb_val["Anchors_Feb'26"],
        ws_formulas=wb_fml["Anchors_Feb'26"],
        header_row=2,
        data_start=3,
        formula_dict=ANCHORS_FORMULAS,
    )

    # ── 2. Anchor-Attendance ───────────────────────────────────────────────
    # Row 1 = group headers (Customer, Project, ..., November Opening leaves, ...)
    # Row 2 = column headers  (S.NO, EMP ID, Names, ..., Total Present, ...)
    # Row 3 = accrual values (0.5, 0.5, 1, 0.25 under leave balance columns)
    # Row 4 = first data row
    write_csv_with_formulas(
        output_path=OUTPUT_DIR / "Anchor_Attendance.csv",
        sheet_name="Anchor-Attendance",
        ws_values=wb_val["Anchor-Attendance"],
        ws_formulas=wb_fml["Anchor-Attendance"],
        header_row=2,
        data_start=4,
        formula_dict=ATTENDANCE_FORMULAS,
    )

    # ── 3. Jan'26- Leave Bal ───────────────────────────────────────────────
    # Row 1 = main header (S.NO, EMP ID, Names, Closing Leave Balance...)
    # Row 2 = sub-headers (Last Year CF, CL, SL, EL, MG, Extra EL)
    # Row 3 = first data row
    write_csv_with_formulas(
        output_path=OUTPUT_DIR / "Jan26_LeaveBal.csv",
        sheet_name="Jan'26- Leave Bal",
        ws_values=wb_val["Jan'26- Leave Bal"],
        ws_formulas=wb_fml["Jan'26- Leave Bal"],
        header_row=2,
        data_start=3,
        formula_dict=LEAVE_BAL_FORMULAS,
    )

    print(f"\nDone! All 3 CSVs written to: {OUTPUT_DIR}")
    print("\nNext step: Review the FORMULA_REFERENCE sections to verify")
    print("the payroll engine matches these formulas exactly.")


if __name__ == "__main__":
    main()
